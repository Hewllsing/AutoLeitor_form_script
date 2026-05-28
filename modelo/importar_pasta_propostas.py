from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import traceback
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
LOCAL_LIBS = SCRIPT_DIR / "libs"
if LOCAL_LIBS.exists():
    sys.path.insert(0, str(LOCAL_LIBS))

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from pypdf import PdfReader


PROJECT_ROOT = SCRIPT_DIR.parent
TEMPLATE_WORKBOOK = SCRIPT_DIR / "Analise comparativa 1.xlsx"
PROPOSALS_DIR = PROJECT_ROOT / "propostas"
RESULTS_DIR = PROJECT_ROOT / "resultados"
OUTPUT_WORKBOOK = RESULTS_DIR / "Analise comparativa preenchida.xlsx"
INTERNAL_DIR = RESULTS_DIR / "nao-mexer"
IMPORTED_LOG = INTERNAL_DIR / "propostas_importadas.json"
REPORT_FILE = INTERNAL_DIR / "resultado_importacao.txt"

SHEET_NAME = "Analise"
DATA_START_ROW = 7
DATA_END_ROW = 14
TEMPLATE_ROW = 12

FIELD_MAP = {
    "processo_numero": "processo",
    "nome_da_empresa_/_company_name": "fornecedor",
    "nif_/_vat_number": "nif",
    "pre\u00e7o_total_/_total_price": "preco",
    "prazo_de_entrega_/_delivery_time": "prazo_entrega",
    "condi\u00e7\u00f5es_de_pagamento_/_payment_terms": "pagamento",
    "garantia_/_warranty": "garantia",
    "transporte_/_transport": "transporte",
    "certifica\u00e7\u00f5es_do_produto_/_product_certifications": "cert_produto",
    "certifica\u00e7\u00f5es_da_empresa_/_company_certifications": "cert_empresa",
    "validade_da_proposta_/_proposal_validity_date": "validade",
}

INPUT_COLUMNS = {
    "fornecedor": "B",
    "preco": "C",
    "transporte": "D",
    "prazo_entrega": "F",
    "pagamento": "G",
    "garantia": "H",
    "cert_produto": "I",
    "cert_empresa": "J",
    "validade": "K",
}

FORMULA_COLUMNS = ("E", "L", "M", "N", "O", "P", "Q", "R")


def clean_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def read_pdf_fields(pdf_path: Path) -> dict[str, Any]:
    reader = PdfReader(str(pdf_path))
    fields = reader.get_fields() or {}
    data: dict[str, Any] = {}

    for pdf_name, internal_name in FIELD_MAP.items():
        raw = fields.get(pdf_name, {}).get("/V", "")
        data[internal_name] = clean_value(raw)

    return data


def parse_number(value: Any) -> float | int | None:
    text = str(value or "").strip()
    if not text:
        return None

    text = text.replace("\u20ac", "").replace(" ", "")
    text = re.sub(r"[^0-9,.\-]", "", text)

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return None

    return int(number) if number.is_integer() else number


def parse_days(value: Any) -> int | None:
    text = str(value or "").strip()
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def normalize_yes_no(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    if text in {"sim", "s", "yes", "y", "true", "1"}:
        return "Sim"
    if text in {"n\u00e3o", "nao", "n", "no", "false", "0"}:
        return "N\u00e3o"
    return str(value).strip()


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_imported_log(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"files": {}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"files": {}}


def save_imported_log(path: Path, log: dict[str, Any]) -> None:
    path.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")


def first_empty_row(ws) -> int | None:
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if not ws[f"B{row}"].value:
            return row
    return None


def copy_template_row(ws, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(TEMPLATE_ROW, col)
        target = ws.cell(target_row, col)

        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)

    for col in FORMULA_COLUMNS:
        source_ref = f"{col}{TEMPLATE_ROW}"
        target_ref = f"{col}{target_row}"
        source_formula = ws[source_ref].value
        if isinstance(source_formula, str) and source_formula.startswith("="):
            ws[target_ref] = Translator(source_formula, origin=source_ref).translate_formula(
                target_ref
            )


def clear_example_rows_if_needed(ws) -> bool:
    example_suppliers = [ws[f"B{row}"].value for row in range(7, 13)]
    if example_suppliers != ["A", "B", "C", "D", "E", "F"]:
        return False

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in INPUT_COLUMNS.values():
            ws[f"{col}{row}"] = None
    return True


def normalized_proposal_values(data: dict[str, Any]) -> dict[str, Any]:
    values = {
        "fornecedor": data["fornecedor"],
        "preco": parse_number(data["preco"]),
        "transporte": parse_number(data["transporte"]) or 0,
        "prazo_entrega": parse_days(data["prazo_entrega"]),
        "pagamento": data["pagamento"],
        "garantia": data["garantia"],
        "cert_produto": normalize_yes_no(data["cert_produto"]),
        "cert_empresa": normalize_yes_no(data["cert_empresa"]),
        "validade": data["validade"],
    }

    missing_required = [
        label
        for label, key in [
            ("nome da empresa", "fornecedor"),
            ("preco", "preco"),
            ("prazo de entrega", "prazo_entrega"),
        ]
        if values[key] in ("", None)
    ]
    if missing_required:
        raise ValueError(
            "Campos essenciais vazios: " + ", ".join(missing_required)
        )

    return values


def write_proposal(ws, row: int, data: dict[str, Any]) -> None:
    values = normalized_proposal_values(data)
    for key, col in INPUT_COLUMNS.items():
        ws[f"{col}{row}"] = values[key]

    processo = data.get("processo")
    if processo:
        ws["C4"] = processo


def load_workbook_for_import(template_path: Path, output_path: Path):
    if output_path.exists():
        return load_workbook(output_path), False

    if not template_path.exists():
        raise FileNotFoundError(f"Planilha modelo nao encontrada: {template_path}")

    workbook = load_workbook(template_path)
    ws = workbook[SHEET_NAME]
    clear_example_rows_if_needed(ws)
    return workbook, True


def import_folder(
    proposals_dir: Path,
    template_path: Path,
    output_path: Path,
    log_path: Path,
) -> tuple[list[str], list[str]]:
    proposals_dir.mkdir(parents=True, exist_ok=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook, created_output = load_workbook_for_import(template_path, output_path)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"A planilha nao tem a aba esperada: {SHEET_NAME}")

    ws = workbook[SHEET_NAME]
    log = load_imported_log(log_path)
    imported_hashes = set(log.get("files", {}).keys())

    imported: list[str] = []
    skipped_or_failed: list[str] = []

    for pdf_path in sorted(proposals_dir.glob("*.pdf")):
        digest = file_hash(pdf_path)
        if digest in imported_hashes:
            continue

        try:
            data = read_pdf_fields(pdf_path)
            row = first_empty_row(ws)
            if row is None:
                raise ValueError(
                    f"Nao ha linha vazia entre {DATA_START_ROW} e {DATA_END_ROW}."
                )

            copy_template_row(ws, row)
            write_proposal(ws, row, data)
            imported.append(f"{pdf_path.name} -> linha {row}")
            log.setdefault("files", {})[digest] = {
                "file_name": pdf_path.name,
                "file_path": str(pdf_path),
                "imported_at": datetime.now().isoformat(timespec="seconds"),
                "row": row,
            }
            imported_hashes.add(digest)
        except Exception as exc:
            skipped_or_failed.append(f"{pdf_path.name}: {exc}")

    if imported or created_output:
        workbook.save(output_path)
        save_imported_log(log_path, log)

    return imported, skipped_or_failed


def write_report(report_path: Path, message: str) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(message, encoding="utf-8")


def build_report(imported: list[str], skipped_or_failed: list[str], output_path: Path) -> str:
    lines = ["Importacao de propostas concluida.", ""]

    if imported:
        lines.append(f"Novas propostas importadas: {len(imported)}")
        lines.extend(f"- {item}" for item in imported)
    else:
        lines.append("Nenhuma proposta nova encontrada.")

    if skipped_or_failed:
        lines.extend(["", "Arquivos nao importados:"])
        lines.extend(f"- {item}" for item in skipped_or_failed)

    lines.extend(["", f"Planilha: {output_path}"])
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Importa todos os PDFs novos de uma pasta para a planilha."
    )
    parser.add_argument("--proposals-dir", type=Path, default=PROPOSALS_DIR)
    parser.add_argument("--template", type=Path, default=TEMPLATE_WORKBOOK)
    parser.add_argument("--output", type=Path, default=OUTPUT_WORKBOOK)
    parser.add_argument("--log", type=Path, default=IMPORTED_LOG)
    parser.add_argument("--report", type=Path, default=REPORT_FILE)
    args = parser.parse_args()

    try:
        imported, skipped_or_failed = import_folder(
            args.proposals_dir.resolve(),
            args.template.resolve(),
            args.output.resolve(),
            args.log.resolve(),
        )
        report = build_report(imported, skipped_or_failed, args.output.resolve())
        write_report(args.report.resolve(), report)
        print(report)
        return 0 if not skipped_or_failed else 1
    except Exception:
        report = "Erro ao importar propostas.\n\n" + traceback.format_exc()
        write_report(args.report.resolve(), report)
        print(report)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
